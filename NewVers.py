import re
import openpyxl
import win32com.client
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By


def find_digit(a):
    num = ''
    for i in list(a):
        if i.isdigit():
            num += i
        elif i == '.':
            num += ','
    return num


# Является ли переменная числом
def isint(a):
    try:
        int(a)
        return True
    except ValueError:
        return False


def seek_price(val):
    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Prices']
    valu = re.sub(r'\s+', ' ', val)
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1, ).value == valu:
            return ws.cell(row=i, column=2).coordinate


def seek_vendor_price(val):
    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Prices']
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1, ).value == val:
            return ws.cell(row=i, column=3).coordinate


# Обновлние HTML
def update_prices():
    driver_path = r'C:\Program Files\Google\Chrome\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path)
    driver.get('https://tarkov-market.com/ru/')
    driver.find_element(By.XPATH, '//div[@class="cell pointer"]').click()
    while True:
        try:
            driver.find_element(By.XPATH, '//span[text()=".300 AAC Blackout AP"]')
        except Exception:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        else:
            break
    names = driver.find_elements(By.XPATH, '//span[@class="name"]')
    pric = driver.find_elements(By.XPATH, '//span[@class="price-main"]')
    vendor = driver.find_elements(By.XPATH, '//div[@class="alt"]')
    titles = []
    prices = []
    vendor_prices = []
    for name in names:
        titles.append(name.text)
    for price in pric:
        prices.append(find_digit(price.text))
    for price in vendor:
        vendor_prices.append(find_digit(price.text))
    try:
        wb = openpyxl.load_workbook('Database.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    try:
        ws = wb['Prices']
    except KeyError:
        ws = wb.active
        ws.title = 'Prices'
    ws.cell(row=1, column=1, value='Name')
    ws.cell(row=1, column=2, value='Price')
    ws.cell(row=1, column=3, value='Vendor price')
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    row = 2
    for i in range(len(names)):
        ws.cell(row, 1).value = titles[i]
        ws.cell(row, 2).value = int(prices[i])
        ws.cell(row, 3).value = int(vendor_prices[i])
        ws.cell(row, 4).value = '=C' + str(row) + '-B' + str(row)
        row += 1
    wb.save('Database.xlsx')
    driver.quit()


def sort():
    # Обновляет
    xlapp = win32com.client.DispatchEx("Excel.Application")
    wb = xlapp.workbooks.open(r'C:\Users\Zina\PycharmProjects\Tarkov-Parser\Database.xlsx')
    wb.RefreshAll()
    wb.Save()
    xlapp.Quit()
    # Сортирует
    try:
        df = pd.read_excel('Database.xlsx', sheet_name='Crafts_raw', engine='openpyxl')
    except ValueError:
        print('Лист "Crafts_raw" не найден, запущена функция "update_crafts"')
        update_crafts()
    sorted_df = df.sort_values(by='Profit/H', ascending=False)
    # Сохраняет сортировку + очищает страницу
    wb = openpyxl.load_workbook('Database.xlsx')
    try:
        ws = wb['Crafts_nude']
    except KeyError:
        ws = wb.create_sheet('Crafts_nude')
    for i in range(2, ws.max_row + 1):
        for y in range(1, ws.max_column + 1):
            ws.cell(row=i, column=y).value = None
    for i in dataframe_to_rows(sorted_df, index=False, header=True):
        ws.append(i)
    wb.save('Database.xlsx')


def update_crafts():
    driver_path = r'C:\Program Files\Google\Chrome\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path)
    driver.get('https://tarkov-market.com/ru/hideout')
    while True:
        try:
            driver.find_element(By.XPATH, '//span[@class="big"][text()="Противогаз ГП-5"]')
        except Exception:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        else:
            break
    cards = driver.find_elements(By.XPATH, '//div[@class="card recipe"]')
    wb = openpyxl.load_workbook('Database.xlsx')
    try:
        ws = wb['Crafts_raw']
    except KeyError:
        ws = wb.create_sheet('Crafts_raw')
        columns = ['Module', 'Ingredient', 'Amount', 'Price', 'Ingredient', 'Amount', 'Price', 'Ingredient', 'Amount',
                   'Price', 'Ingredient', 'Amount', 'Price', 'Ingredient', 'Amount', 'Price', 'Sum', 'Time(min)',
                   'Name', 'Amount', 'Price', 'Sum', 'Profit', 'Profit/H']
        for i in range(1, len(columns) + 1):
            ws.cell(1, i, value=columns[i - 1])
    row = 2
    for i in range(1, len(cards) + 1):
        column = 1
        ingredients = []
        in_amount = []
        pricescord = []
        names = driver.find_elements(By.XPATH, '//div[@class="card recipe"][' + str(i) + ']//span[@class="big"]')
        for y in range(1, len(names)):
            ingredients.append(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                                   ']//div[@class="d-flex only mb-15"][' + str(y) +
                                                   ']//span').get_attribute('textContent'))
            in_amount.append(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                                 ']//div[@class="d-flex only mb-15"][' + str(y) +
                                                 ']//div[@class="image"]/div').get_attribute('textContent'))
            pricescord.append(seek_price(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                                             ']//div[@class="d-flex only mb-15"][' + str(y) +
                                                             ']//span').get_attribute('textContent')))
        modules = driver.find_element(By.XPATH, "//div[@class='row recipe'][" + str(i) + "]//div[@class='big']").text
        result = driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                     ']//div[@class="d-flex only mb-15"][' + str(len(names)) +
                                     ']//span').get_attribute('textContent')
        time = list(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                        ']//div[@class="text-center big"]').get_attribute('textContent'))
        minutes = 0
        tim = ''
        for t in range(len(time)):
            if isint(time[t]):
                tim += time[t]
            elif time[t] == 'ч':
                minutes = int(tim) * 60
                tim = ''
            elif time[t] == 'м':
                minutes += int(tim)
                break
        result_amount = driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                            ']//div[@class="d-flex only mb-15"][' + str(len(names)) +
                                            ']//div[@class="image"]/div').get_attribute('textContent')
        result_price_coordinate = seek_price(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                                                 ']//div[@class="d-flex only mb-15"][' + str(
            len(names)) +
                                                                 ']//span').get_attribute('textContent'))
        ws.cell(row=row, column=column, value=modules)
        column += 1
        for y in range(1, 6):
            try:
                ws.cell(row=row, column=column, value=ingredients[y - 1])
                column += 1
                ws.cell(row=row, column=column, value=find_digit(in_amount[y - 1]))
                column += 1
                ws.cell(row=row, column=column, value='=Prices!' + pricescord[y - 1])
                column += 1
            except IndexError:
                column += 3
        ws.cell(row=row, column=column,
                value=('=D' + str(row) + '*C' + str(row) + '+G' + str(row) + '*F' + str(row) + '+J' + str(row) +
                       '*I' + str(row) + '+M' + str(row) + '*L' + str(row) + '+P' + str(row) + '*O' + str(row)))
        column += 1
        ws.cell(row=row, column=column, value=minutes)
        column += 1
        ws.cell(row=row, column=column, value=result)
        column += 1
        ws.cell(row=row, column=column, value=find_digit(result_amount))
        column += 1
        ws.cell(row=row, column=column, value='=Prices!' + str(result_price_coordinate))
        column += 1
        ws.cell(row=row, column=column, value='=U' + str(row) + '*T' + str(row))
        column += 1
        ws.cell(row=row, column=column, value='=V' + str(row) + '-Q' + str(row))
        column += 1
        ws.cell(row=row, column=column, value='=W' + str(row) + '/R' + str(row) + '*60')
        row += 1
    wb.save('Database.xlsx')
    driver.quit()


def make_table():
    wb = openpyxl.load_workbook('Database.xlsx')
    try:
        ws = wb['Crafts']
    except KeyError:
        ws = wb.create_sheet('Crafts')
        columns = ['Module', 'Ingredients', 'Amount', 'Price', 'Sum', 'Time(min)',
                   'Name', 'Amount', 'Price', 'Sum', 'Profit', 'Profit/H']
        for i in range(1, len(columns) + 1):
            ws.cell(1, i, value=columns[i - 1])
    row = 2
    for i in range(124):
        ws.cell(row=row, column=1, value='=Crafts_nude!A' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=1)
        for y in range(0, 5):
            ws.cell(row=row + y, column=2, value='=Crafts_nude!' + str(chr(ord('B') + (y * 3))) + str(i + 2))
            ws.cell(row=row + y, column=3, value='=Crafts_nude!' + str(chr(ord('C') + (y * 3))) + str(i + 2))
            ws.cell(row=row + y, column=4, value='=Crafts_nude!' + str(chr(ord('D') + (y * 3))) + str(i + 2))
        ws.cell(row=row, column=5, value='=Crafts_nude!Q' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=5, end_row=row + 4, end_column=5)
        ws.cell(row=row, column=6, value='=Crafts_nude!R' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=6, end_row=row + 4, end_column=6)
        ws.cell(row=row, column=7, value='=Crafts_nude!S' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=7, end_row=row + 4, end_column=7)
        ws.cell(row=row, column=8, value='=Crafts_nude!T' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=8, end_row=row + 4, end_column=8)
        ws.cell(row=row, column=9, value='=Crafts_nude!U' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=9, end_row=row + 4, end_column=9)
        ws.cell(row=row, column=10, value='=I' + str(row) + '*H' + str(row))
        ws.merge_cells(start_row=row, start_column=10, end_row=row + 4, end_column=10)
        ws.cell(row=row, column=11, value='=J' + str(row) + '-E' + str(row))
        ws.merge_cells(start_row=row, start_column=11, end_row=row + 4, end_column=11)
        ws.cell(row=row, column=12, value='=K' + str(row) + '/F' + str(row) + '*60')
        ws.merge_cells(start_row=row, start_column=12, end_row=row + 4, end_column=12)
        row += 5
    wb.save('Database.xlsx')


def update_barters():
    driver_path = r'C:\Program Files\Google\Chrome\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path)
    driver.get('https://tarkov-market.com/ru/barter')
    while True:
        try:
            driver.find_element(By.XPATH, '//span[@class="big"][text()="Бронежилет 6Б43 6А Забрало-Ш (0/85)"]')
        except Exception:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        else:
            break
    cards = driver.find_elements(By.XPATH, '//div[@class="card recipe"]')
    wb = openpyxl.load_workbook('Database.xlsx')
    try:
        ws = wb['Barters_nude']
    except KeyError:
        ws = wb.create_sheet('Barters_nude')
        columns = ['Module', 'Ingredient', 'Amount', 'Price', 'Ingredient', 'Amount', 'Price', 'Ingredient', 'Amount',
                   'Price', 'Ingredient', 'Amount', 'Price', 'Ingredient', 'Amount', 'Price', 'Sum', 'Name', 'Amount',
                   'Price', 'Sum', 'Profit']
        for i in range(1, len(columns) + 1):
            ws.cell(1, i, value=columns[i - 1])
    row = 2
    for i in range(1, len(cards) + 1):
        ingredients = []
        in_amount = []
        prices_coordinates = []
        trader = driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(i) +
                                     ']//div[@class="big"]').get_attribute('textContent')
        names = driver.find_elements(By.XPATH, '//div[@class="card recipe"][' + str(i) + ']//span[@class="big"]')
        for y in range(1, len(names)):
            ingredients.append(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(
                i) + ']//div[@class="d-flex only mb-15"][' + str(y) + ']//span').get_attribute('textContent'))
            in_amount.append(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(
                i) + ']//div[@class="d-flex only mb-15"][' + str(y) + ']//div[@class="image"]/div').get_attribute(
                'textContent'))
            prices_coordinates.append(seek_price(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(
                i) + ']//div[@class="d-flex only mb-15"][' + str(y) + ']//span').get_attribute('textContent')))
        result = driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(
            i) + ']//div[@class="d-flex only mb-15"][' + str(len(names)) + ']//span').get_attribute('textContent')
        result_amount = driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(
            i) + ']//div[@class="d-flex only mb-15"][' + str(len(names)) + ']//div[@class="image"]/div').get_attribute(
            'textContent')
        result_price_coordinate = seek_price(driver.find_element(By.XPATH, '//div[@class="card recipe"][' + str(
            i) + ']//div[@class="d-flex only mb-15"][' + str(len(names)) + ']//span').get_attribute('textContent'))
        column = 1
        ws.cell(row=row, column=column, value=trader)
        column += 1
        for y in range(1, 6):
            try:
                ws.cell(row=row, column=column, value=ingredients[y - 1])
                column += 1
                ws.cell(row=row, column=column, value=find_digit(in_amount[y - 1]))
                column += 1
                ws.cell(row=row, column=column, value='=Prices!' + prices_coordinates[y - 1])
                column += 1
            except IndexError:
                column += 3
        ws.cell(row=row, column=column, value=(
                '=D' + str(row) + '*C' + str(row) + '+G' + str(row) + '*F' + str(row) + '+J' + str(row) +
                '*I' + str(row) + '+M' + str(row) + '*L' + str(row) + '+P' + str(row) + '*O' + str(row)))
        column += 1
        ws.cell(row=row, column=column, value=result)
        column += 1
        ws.cell(row=row, column=column, value=find_digit(result_amount))
        column += 1
        ws.cell(row=row, column=column, value='=Prices!' + str(result_price_coordinate))
        column += 1
        ws.cell(row=row, column=column, value='=S' + str(row) + '*T' + str(row))
        column += 1
        ws.cell(row=row, column=column, value='=U' + str(row) + '-Q' + str(row))
        row += 1
    wb.save('Database.xlsx')
    driver.quit()


def make_barters_table():
    wb = openpyxl.load_workbook('Database.xlsx')
    try:
        ws = wb['Barters']
    except KeyError:
        ws = wb.create_sheet('Barters')
        columns = ['Module', 'Ingredients', 'Amount', 'Price', 'Sum', 'Name', 'Amount', 'Price', 'Sum', 'Profit']
        for i in range(1, len(columns) + 1):
            ws.cell(1, i, value=columns[i - 1])
    row = 2
    for i in range(124):
        ws.cell(row=row, column=1, value='=Barters_nude!A' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=1)
        for y in range(0, 5):
            ws.cell(row=row + y, column=2, value='=Barters_nude!' + str(chr(ord('B') + (y * 3))) + str(i + 2))
            ws.cell(row=row + y, column=3, value='=Barters_nude!' + str(chr(ord('C') + (y * 3))) + str(i + 2))
            ws.cell(row=row + y, column=4, value='=Barters_nude!' + str(chr(ord('D') + (y * 3))) + str(i + 2))
        ws.cell(row=row, column=5, value='=Barters_nude!Q' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=5, end_row=row + 4, end_column=5)
        ws.cell(row=row, column=6, value='=Barters_nude!R' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=6, end_row=row + 4, end_column=6)
        ws.cell(row=row, column=7, value='=Barters_nude!S' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=7, end_row=row + 4, end_column=7)
        ws.cell(row=row, column=8, value='=Barters_nude!T' + str(i + 2))
        ws.merge_cells(start_row=row, start_column=8, end_row=row + 4, end_column=8)
        ws.cell(row=row, column=9, value='=G' + str(row) + '*H' + str(row))
        ws.merge_cells(start_row=row, start_column=9, end_row=row + 4, end_column=9)
        ws.cell(row=row, column=10, value='=I' + str(row) + '-E' + str(row))
        ws.merge_cells(start_row=row, start_column=10, end_row=row + 4, end_column=11)
        row += 5
    wb.save('Database.xlsx')


if __name__ == '__main__':
    make_barters_table()

# TODO: Попробовать новенькое:
#   синхронный код
#   Попробовать оптимизировать (https://medium.com/nuances-of-programming/как-ускорить-python-8df43f87ef6f)
#   Перевести браузер в headless или хоть спрятать его
# TODO: Доавить в таблицу:
#   динамику цены
#   графики изменения цены
# TODO: Предотвращать возможные ошибки
# TODO: Попробовать исправить ошибку с двойными крафтами
# TODO: научиться избавляться от двойных пробелов в названиях придмета (seek_price)
